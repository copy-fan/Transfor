#!/usr/bin/env python3
"""Convert an Excel workbook into a Feishu/Lark Base via lark-cli."""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any

VALID_HUES = [
    "Blue",
    "Green",
    "Orange",
    "Purple",
    "Red",
    "Yellow",
    "Wathet",
    "Lime",
    "Turquoise",
    "Carmine",
    "Gray",
]

MAIN_SHEET_HINTS = ("飞书导入", "导入主表", "主表", "明细", "数据", "raw", "main")
ID_RE = re.compile(r"(?:^|[ _-])(id|编号|编码|token|素材id|订单号|单号)(?:$|[ _-])", re.I)
PERCENT_RE = re.compile(r"(率|占比|ctr|cvr|rate|percent|percentage)", re.I)
CURRENCY_RE = re.compile(r"(金额|消耗|花费|成本|gmv|收入|营收|销售额|revenue|amount|cost|spend)", re.I)
DATETIME_RE = re.compile(r"(时间|日期|date|time|created|updated)", re.I)
TAG_RE = re.compile(r"(标签|tag|评估)", re.I)


def clean_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        if value in ("", "-"):
            return None
        return value
    return value


def split_multi(value: Any) -> list[str] | None:
    value = clean_value(value)
    if not value:
        return None
    parts = [p.strip() for p in re.split(r"[、，,;/|]", str(value)) if p.strip() and p.strip() != "-"]
    return parts or None


def unique(values: list[Any], limit: int | None = None) -> list[Any]:
    out: list[Any] = []
    for value in values:
        value = clean_value(value)
        if value is None:
            continue
        if value not in out:
            out.append(value)
            if limit and len(out) >= limit:
                break
    return out


def option(name: Any, idx: int) -> dict[str, str]:
    return {"name": str(name), "hue": VALID_HUES[idx % len(VALID_HUES)], "lightness": "Lighter"}


def normalize_header(value: Any, idx: int) -> str:
    value = clean_value(value)
    return str(value) if value else f"列{idx}"


def sheet_rows(ws: Any) -> tuple[list[str], list[list[Any]]]:
    headers = [normalize_header(ws.cell(1, c).value, c) for c in range(1, ws.max_column + 1)]
    while headers and all(clean_value(ws.cell(r, len(headers)).value) is None for r in range(1, ws.max_row + 1)):
        headers.pop()
    rows: list[list[Any]] = []
    for r in range(2, ws.max_row + 1):
        row = [clean_value(ws.cell(r, c).value) for c in range(1, len(headers) + 1)]
        if any(v is not None for v in row):
            rows.append(row)
    return headers, rows


def choose_primary_sheet(wb: Any, requested: str | None) -> str:
    if requested:
        if requested not in wb.sheetnames:
            raise SystemExit(f"Primary sheet not found: {requested}")
        return requested
    non_empty = []
    for ws in wb.worksheets:
        headers, rows = sheet_rows(ws)
        if headers and rows:
            non_empty.append((ws.title, len(rows), len(headers)))
    for hint in MAIN_SHEET_HINTS:
        for name, _, _ in non_empty:
            if hint.lower() in name.lower():
                return name
    if not non_empty:
        raise SystemExit("Workbook has no non-empty sheets.")
    return max(non_empty, key=lambda x: x[1] * x[2])[0]


def infer_field(name: str, values: list[Any], primary: bool = False) -> dict[str, Any]:
    non_empty = [clean_value(v) for v in values if clean_value(v) is not None]
    lower = name.lower()
    if primary or ID_RE.search(name):
        return {"type": "text", "name": name}
    if DATETIME_RE.search(name):
        return {"type": "datetime", "name": name, "style": {"format": "yyyy-MM-dd HH:mm"}}
    numeric_count = sum(isinstance(v, (int, float)) and not isinstance(v, bool) for v in non_empty)
    if non_empty and numeric_count / len(non_empty) >= 0.85:
        if CURRENCY_RE.search(name):
            return {"type": "number", "name": name, "style": {"type": "currency", "precision": 2, "currency_code": "CNY"}}
        if PERCENT_RE.search(name):
            return {
                "type": "number",
                "name": name,
                "style": {"type": "plain", "precision": 2, "percentage": True, "thousands_separator": False},
            }
        precision = 0 if all(float(v).is_integer() for v in non_empty if isinstance(v, (int, float))) else 2
        return {
            "type": "number",
            "name": name,
            "style": {"type": "plain", "precision": precision, "percentage": False, "thousands_separator": precision == 0},
        }
    uniq = unique(non_empty)
    if 0 < len(uniq) <= min(50, max(12, len(non_empty) // 2)):
        if TAG_RE.search(name):
            opts = unique([p for v in non_empty for p in (split_multi(v) or [])])
            return {"type": "select", "name": name, "multiple": True, "options": [option(v, i) for i, v in enumerate(opts)]}
        return {"type": "select", "name": name, "multiple": False, "options": [option(v, i) for i, v in enumerate(uniq)]}
    return {"type": "text", "name": name}


def cell_for_field(field: dict[str, Any], value: Any) -> Any:
    value = clean_value(value)
    if value is None:
        return None
    if field["type"] == "select" and field.get("multiple"):
        return split_multi(value)
    if field["type"] == "datetime":
        return str(value)
    if field["type"] == "number" and isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    return value


def build_table_payload(headers: list[str], rows: list[list[Any]], force_text: bool = False) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    columns = list(zip(*rows)) if rows else [[] for _ in headers]
    fields: list[dict[str, Any]] = []
    for idx, header in enumerate(headers):
        values = list(columns[idx]) if idx < len(columns) else []
        fields.append({"type": "text", "name": header} if force_text else infer_field(header, values, primary=idx == 0))
    out_rows = [[cell_for_field(fields[i], value) for i, value in enumerate(row)] for row in rows]
    return fields, {"fields": headers, "rows": out_rows}


def find_field(headers: list[str], patterns: tuple[str, ...]) -> str | None:
    for pattern in patterns:
        regex = re.compile(pattern, re.I)
        for header in headers:
            if regex.search(header):
                return header
    return None


def run_cli(args: list[str], cwd: Path, log_path: Path) -> dict[str, Any]:
    proc = subprocess.run(args, cwd=cwd, text=True, capture_output=True)
    log_path.write_text(proc.stdout + ("\nSTDERR:\n" + proc.stderr if proc.stderr else ""), encoding="utf-8")
    try:
        data = json.loads(proc.stdout)
    except Exception as exc:
        raise SystemExit(f"Command did not return JSON: {' '.join(args)}\n{proc.stdout}\n{proc.stderr}") from exc
    if not data.get("ok"):
        raise SystemExit(json.dumps(data, ensure_ascii=False, indent=2))
    return data


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")


def create_views(cli: str, cwd: Path, base_token: str, table_id: str, headers: list[str]) -> list[dict[str, Any]]:
    spend = find_field(headers, (r"消耗|花费|成本|spend|cost",))
    roi = find_field(headers, (r"^roi$|支付roi",))
    script = find_field(headers, (r"脚本风格|script",))
    hook = find_field(headers, (r"开头钩子|hook",))
    tag = find_field(headers, (r"原始标签|标签|tag",))
    value = find_field(headers, (r"素材值|value|score",)) or spend
    views: list[dict[str, Any]] = []
    specs: list[dict[str, Any]] = []
    if roi and spend:
        specs.append({
            "name": "08_ROI消耗高潜池",
            "filter": {"logic": "and", "conditions": [[roi, ">=", 2], [spend, ">=", 50]]},
            "sort": {"sort_config": [{"field": value, "desc": True}, {"field": spend, "desc": True}]},
        })
        specs.append({
            "name": "09_低ROI高消耗预警",
            "filter": {"logic": "and", "conditions": [[spend, ">=", 100], [roi, "<", 1]]},
            "sort": {"sort_config": [{"field": spend, "desc": True}]},
        })
    if spend and (script or hook or tag):
        conditions = []
        if script:
            conditions.append([script, "intersects", ["待补充"]])
        if hook:
            conditions.append([hook, "intersects", ["待补充"]])
        if tag:
            conditions.append([tag, "empty"])
        specs.append({
            "name": "10_标签补齐优先队列",
            "filter": {"logic": "or", "conditions": conditions},
            "sort": {"sort_config": [{"field": spend, "desc": True}]},
        })
    for spec in specs:
        created = run_cli(
            [cli, "base", "+view-create", "--base-token", base_token, "--table-id", table_id, "--json", json.dumps({"name": spec["name"], "type": "grid"}, ensure_ascii=False), "--format", "json"],
            cwd,
            cwd / f"view_{spec['name']}_create.json",
        )
        data = created["data"]
        view = data.get("view") or (data.get("views") or [None])[0]
        vid = view["id"]
        run_cli([cli, "base", "+view-set-filter", "--base-token", base_token, "--table-id", table_id, "--view-id", vid, "--json", json.dumps(spec["filter"], ensure_ascii=False), "--format", "json"], cwd, cwd / f"view_{spec['name']}_filter.json")
        run_cli([cli, "base", "+view-set-sort", "--base-token", base_token, "--table-id", table_id, "--view-id", vid, "--json", json.dumps(spec["sort"], ensure_ascii=False), "--format", "json"], cwd, cwd / f"view_{spec['name']}_sort.json")
        views.append({"name": spec["name"], "view_id": vid})
    return views


def create_dashboards(cli: str, cwd: Path, base_token: str, table_name: str, headers: list[str]) -> list[dict[str, Any]]:
    spend = find_field(headers, (r"消耗|花费|成本|spend|cost",))
    script = find_field(headers, (r"脚本风格|script",))
    hook = find_field(headers, (r"开头钩子|hook",))
    if not spend or not (script or hook):
        return []
    dash = run_cli([cli, "base", "+dashboard-create", "--base-token", base_token, "--name", "素材消耗维度饼图", "--format", "json"], cwd, cwd / "dashboard_create.json")
    dashboard_id = dash["data"]["dashboard"]["dashboard_id"]
    blocks = []
    for dim in [script, hook]:
        if not dim:
            continue
        config = {
            "table_name": table_name,
            "series": [{"field_name": spend, "rollup": "SUM"}],
            "group_by": [{"field_name": dim, "mode": "integrated", "sort": {"type": "value", "order": "desc"}}],
        }
        block = run_cli(
            [cli, "base", "+dashboard-block-create", "--base-token", base_token, "--dashboard-id", dashboard_id, "--name", f"{dim}消耗占比", "--type", "pie", "--data-config", json.dumps(config, ensure_ascii=False), "--format", "json"],
            cwd,
            cwd / f"dashboard_block_{dim}.json",
        )
        blocks.append({"name": f"{dim}消耗占比", "block_id": block["data"]["block"]["block_id"]})
    if blocks:
        run_cli([cli, "base", "+dashboard-arrange", "--base-token", base_token, "--dashboard-id", dashboard_id, "--format", "json"], cwd, cwd / "dashboard_arrange.json")
    return [{"name": "素材消耗维度饼图", "dashboard_id": dashboard_id, "blocks": blocks}]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", required=True, type=Path)
    parser.add_argument("--base-name")
    parser.add_argument("--primary-sheet")
    parser.add_argument("--work-dir", type=Path)
    parser.add_argument("--lark-cli", default=shutil.which("lark-cli") or "/Users/cp/.local/bin/lark-cli")
    parser.add_argument("--prepare-only", action="store_true")
    parser.add_argument("--no-views", action="store_true")
    parser.add_argument("--no-dashboards", action="store_true")
    args = parser.parse_args()

    if not args.excel.exists():
        raise SystemExit(f"Excel file not found: {args.excel}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover
        raise SystemExit(
            "openpyxl is required. Run with the bundled Python from load_workspace_dependencies."
        ) from exc
    wb = load_workbook(args.excel, data_only=True)
    primary = choose_primary_sheet(wb, args.primary_sheet)
    base_name = args.base_name or args.excel.stem
    work_dir = args.work_dir or Path(tempfile.mkdtemp(prefix="feishu-base-import-"))
    work_dir.mkdir(parents=True, exist_ok=True)

    ws = wb[primary]
    headers, rows = sheet_rows(ws)
    fields, records = build_table_payload(headers, rows)
    write_json(work_dir / "main_fields.json", fields)
    for idx in range(0, len(records["rows"]), 200):
        write_json(work_dir / f"main_records_{idx // 200 + 1}.json", {"fields": records["fields"], "rows": records["rows"][idx : idx + 200]})

    auxiliaries = []
    for sheet in wb.sheetnames:
        if sheet == primary:
            continue
        h, r = sheet_rows(wb[sheet])
        if not h or not r:
            continue
        aux_fields, aux_records = build_table_payload(h, r, force_text=("原始" in sheet))
        safe = re.sub(r"[^\w\u4e00-\u9fff-]+", "_", sheet)
        write_json(work_dir / f"{safe}_fields.json", aux_fields)
        write_json(work_dir / f"{safe}_records.json", aux_records)
        auxiliaries.append({"name": sheet, "fields_file": f"{safe}_fields.json", "records_file": f"{safe}_records.json", "rows": len(r)})

    plan = {"base_name": base_name, "primary_sheet": primary, "main_rows": len(rows), "main_fields": len(headers), "auxiliary_tables": auxiliaries, "work_dir": str(work_dir)}
    write_json(work_dir / "plan.json", plan)
    if args.prepare_only:
        print(json.dumps({"ok": True, "plan": plan, "fields": fields}, ensure_ascii=False, indent=2))
        return

    if not Path(args.lark_cli).exists() and not shutil.which(args.lark_cli):
        raise SystemExit(f"lark-cli not found: {args.lark_cli}")

    create = run_cli(
        [args.lark_cli, "base", "+base-create", "--name", base_name, "--time-zone", "Asia/Shanghai", "--table-name", primary, "--fields", json.dumps(fields, ensure_ascii=False), "--format", "json"],
        work_dir,
        work_dir / "create_base.json",
    )
    base_token = create["data"]["base"]["base_token"]
    table_id = create["data"]["table"]["id"]
    for file in sorted(work_dir.glob("main_records_*.json")):
        run_cli([args.lark_cli, "base", "+record-batch-create", "--base-token", base_token, "--table-id", table_id, "--json", f"@{file.name}", "--format", "json"], work_dir, work_dir / f"{file.stem}_result.json")

    aux_results = []
    for aux in auxiliaries:
        fjson = (work_dir / aux["fields_file"]).read_text(encoding="utf-8")
        table = run_cli([args.lark_cli, "base", "+table-create", "--base-token", base_token, "--name", aux["name"], "--fields", fjson, "--format", "json"], work_dir, work_dir / f"{aux['name']}_create.json")
        aux_table_id = table["data"]["table"]["id"]
        run_cli([args.lark_cli, "base", "+record-batch-create", "--base-token", base_token, "--table-id", aux_table_id, "--json", f"@{aux['records_file']}", "--format", "json"], work_dir, work_dir / f"{aux['name']}_records_result.json")
        aux_results.append({"name": aux["name"], "table_id": aux_table_id, "rows": aux["rows"]})

    views = [] if args.no_views else create_views(args.lark_cli, work_dir, base_token, table_id, headers)
    dashboards = [] if args.no_dashboards else create_dashboards(args.lark_cli, work_dir, base_token, primary, headers)
    summary = {
        "ok": True,
        "base_name": base_name,
        "base_token": base_token,
        "url": create["data"]["base"]["url"],
        "main_table": {"name": primary, "table_id": table_id, "rows": len(rows)},
        "auxiliary_tables": aux_results,
        "views": views,
        "dashboards": dashboards,
        "work_dir": str(work_dir),
    }
    write_json(work_dir / "summary.json", summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
